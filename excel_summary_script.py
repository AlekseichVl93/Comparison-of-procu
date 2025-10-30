
import os
import tempfile
from flask import Flask, request, render_template_string, send_file, flash, redirect, url_for
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

def apply_borders_to_range(worksheet, start_row, start_col, end_row, end_col):
    """
    Применяет границы ко всем ячейкам в указанном диапазоне.
    
    Args:
        worksheet: Лист Excel для форматирования
        start_row: Начальная строка (1-based)
        start_col: Начальный столбец (1-based)
        end_row: Конечная строка (1-based)
        end_col: Конечный столбец (1-based)
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            worksheet.cell(row=row, column=col).border = thin_border

def apply_currency_format_to_price_columns(worksheet, headers_row_2, start_data_row, end_data_row):
    """
    Применяет валютный формат к столбцам с ценами.
    
    Args:
        worksheet: Лист Excel для форматирования
        headers_row_2: Список заголовков второй строки
        start_data_row: Начальная строка с данными (1-based)
        end_data_row: Конечная строка с данными (1-based)
    """
    currency_format = '#,##0.00 ₽'
    
    # Находим столбцы с ценами
    for col_idx, header in enumerate(headers_row_2, start=1):
        if header == 'Цена без НДС за шт':
            # Применяем валютный формат к столбцу с ценами
            for row in range(start_data_row, end_data_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = currency_format

def highlight_minimum_prices(worksheet, headers_row_2, start_data_row, end_data_row):
    """
    Выделяет минимальные цены зеленым цветом в каждой строке.
    
    Args:
        worksheet: Лист Excel для форматирования
        headers_row_2: Список заголовков второй строки
        start_data_row: Начальная строка с данными (1-based)
        end_data_row: Конечная строка с данными (1-based)
    """
    # Находим индексы столбцов с ценами
    price_columns = []
    for col_idx, header in enumerate(headers_row_2, start=1):
        if header == 'Цена без НДС за шт':
            price_columns.append(col_idx)
    
    if not price_columns:
        return
    
    green_font = Font(color='008000')  # Зеленый цвет
    
    # Проходим по каждой строке с данными
    for row in range(start_data_row, end_data_row + 1):
        prices = []
        valid_cells = []
        
        # Собираем цены из всех столбцов с ценами в текущей строке
        for col in price_columns:
            cell = worksheet.cell(row=row, column=col)
            if cell.value is not None:
                try:
                    price = float(cell.value)
                    prices.append(price)
                    valid_cells.append(cell)
                except (ValueError, TypeError):
                    # Пропускаем ячейки с некорректными значениями
                    continue
        
        # Если есть валидные цены, находим минимальную
        if prices:
            min_price = min(prices)
            
            # Выделяем все ячейки с минимальной ценой зеленым цветом
            for i, price in enumerate(prices):
                if price == min_price:
                    valid_cells[i].font = green_font

def set_column_widths_and_wrap_text(worksheet, headers_row_2):
    """
    Устанавливает ширину колонок и включает перенос текста для всех ячеек.
    
    Args:
        worksheet: Лист Excel для форматирования
        headers_row_2: Список заголовков второй строки
    """
    # Устанавливаем ширину колонки A (Наименование)
    worksheet.column_dimensions['A'].width = 63.45
    
    # Словарь с шириной колонок для разных типов заголовков
    column_widths = {
        'Количество предложенное': 10.27,
        'Комментарий поставщика': 22.73,
        'Сроки поставки': 8.45,
        'Цена без НДС за шт': 17.18
    }
    
    # Устанавливаем ширину колонок
    for col_idx, header in enumerate(headers_row_2, start=1):
        if header in column_widths:
            # Получаем букву колонки из ячейки второй строки (не объединенной)
            column_letter = worksheet.cell(row=2, column=col_idx).column_letter
            worksheet.column_dimensions[column_letter].width = column_widths[header]
    
    # Включаем перенос текста для всех ячеек
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.alignment:
                # Сохраняем существующее выравнивание и добавляем перенос текста
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=True
                )
            else:
                # Устанавливаем только перенос текста
                cell.alignment = Alignment(wrap_text=True)

def format_header_rows(worksheet, max_col):
    """
    Форматирует строки заголовков (1 и 2) - жирные внешние границы вокруг всего блока и светло-голубая заливка.
    
    Args:
        worksheet: Лист Excel для форматирования
        max_col: Максимальное количество колонок
    """
    print(f"Форматируем заголовки: строки 1-2, колонки 1-{max_col}")
    
    # Светло-голубая заливка
    light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    
    # Применяем заливку ко всем ячейкам в блоке заголовков
    for row in range(1, 3):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = light_blue_fill
            print(f"  Заливка применена к ячейке {cell.coordinate}")
    
    # Применяем жирные внешние границы к блоку заголовков (строки 1-2, колонки A до max_col)
    apply_thick_borders_to_group(worksheet, 1, 2, max_col)
    print("✓ Форматирование заголовков завершено")

def format_main_product_groups(worksheet, summary_rows, max_col, start_row=3):
    """
    Форматирует группы товаров - жирные границы вокруг каждой группы и увеличенный шрифт для основных товаров.
    Определяет группы по товарам с жирным шрифтом (основные товары) и их последующим аналогам/вариантам.
    
    Args:
        worksheet: Лист Excel для форматирования
        summary_rows: Список строк сводной таблицы
        max_col: Максимальное количество колонок
        start_row: Начальная строка данных (по умолчанию 3)
    """
    # Увеличенный жирный шрифт для основных товаров (на 2 больше базового)
    main_product_font = Font(size=13, bold=True)  # Базовый 11 + 2 = 13
    
    print(f"Начинаем форматирование групп товаров. Всего строк: {len(summary_rows)}")
    
    # ЭТАП 1: Сначала применяем жирный шрифт к основным товарам и определяем их позиции
    main_product_positions = []
    
    for i, row_data in enumerate(summary_rows):
        # Определяем, является ли это основным товаром (не содержит "вариант" или "аналог")
        is_main_product = not ('(вариант' in row_data['name'] or '(аналог' in row_data['name'])
        
        if is_main_product:
            # Применяем увеличенный жирный шрифт к основным товарам
            worksheet.cell(row=start_row + i, column=1).font = main_product_font
            main_product_positions.append(i)
            print(f"Найден основной товар в строке {start_row + i}: '{row_data['name'][:50]}...'")
    
    print(f"Найдено основных товаров: {len(main_product_positions)} в позициях: {main_product_positions}")
    
    # ЭТАП 2: Определяем группы на основе позиций основных товаров
    groups = []
    
    for i, main_pos in enumerate(main_product_positions):
        group_start = main_pos
        
        # Определяем конец группы: до следующего основного товара или до конца списка
        if i + 1 < len(main_product_positions):
            group_end = main_product_positions[i + 1] - 1  # До следующего основного товара
        else:
            group_end = len(summary_rows) - 1  # До конца списка
        
        # Добавляем группу
        group = {
            'start_row': start_row + group_start,
            'end_row': start_row + group_end,
            'main_product': summary_rows[main_pos]['name']
        }
        groups.append(group)
        
        print(f"Создана группа {i+1}: строки {group['start_row']}-{group['end_row']} для товара '{group['main_product'][:50]}...'")
    
    # ЭТАП 3: Применяем жирные границы к каждой группе
    for i, group in enumerate(groups):
        print(f"\nПрименяем жирные границы к группе {i+1}:")
        apply_thick_borders_to_group(worksheet, group['start_row'], group['end_row'], max_col)
        print(f"✓ Применены жирные границы к группе: строки {group['start_row']}-{group['end_row']}")
    
    print(f"\n✓ Форматирование групп завершено. Обработано групп: {len(groups)}")

def apply_thick_borders_to_group(worksheet, start_row, end_row, max_col):
    """
    Применяет жирные внешние границы вокруг всей группы как единого блока.
    
    Args:
        worksheet: Лист Excel для форматирования
        start_row: Начальная строка группы
        end_row: Конечная строка группы
        max_col: Максимальное количество колонок
    """
    print(f"Применяем жирные внешние границы к группе: строки {start_row}-{end_row}, колонки 1-{max_col}")
    
    # Используем разные варианты стиля для совместимости
    thick_side = Side(style='thick', color='000000')
    
    # Верхняя граница группы (вся первая строка)
    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=start_row, column=col)
        # ИСПРАВЛЕНИЕ: Всегда создаем новый объект Border
        old_border = cell.border
        cell.border = Border(
            left=old_border.left if old_border and old_border.left else Side(style='thin'),
            right=old_border.right if old_border and old_border.right else Side(style='thin'),
            top=thick_side,  # Жирная верхняя граница
            bottom=old_border.bottom if old_border and old_border.bottom else Side(style='thin')
        )
        print(f"    Верхняя граница: ячейка {cell.coordinate}")
    
    # Нижняя граница группы (вся последняя строка)
    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=end_row, column=col)
        old_border = cell.border
        cell.border = Border(
            left=old_border.left if old_border and old_border.left else Side(style='thin'),
            right=old_border.right if old_border and old_border.right else Side(style='thin'),
            top=old_border.top if old_border and old_border.top else Side(style='thin'),
            bottom=thick_side  # Жирная нижняя граница
        )
        print(f"    Нижняя граница: ячейка {cell.coordinate}")
    
    # Левая граница группы (весь первый столбец)
    for row in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row, column=1)
        old_border = cell.border
        cell.border = Border(
            left=thick_side,  # Жирная левая граница
            right=old_border.right if old_border and old_border.right else Side(style='thin'),
            top=old_border.top if old_border and old_border.top else Side(style='thin'),
            bottom=old_border.bottom if old_border and old_border.bottom else Side(style='thin')
        )
        print(f"    Левая граница: ячейка {cell.coordinate}")
    
    # Правая граница группы (весь последний столбец)
    for row in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row, column=max_col)
        old_border = cell.border
        cell.border = Border(
            left=old_border.left if old_border and old_border.left else Side(style='thin'),
            right=thick_side,  # Жирная правая граница
            top=old_border.top if old_border and old_border.top else Side(style='thin'),
            bottom=old_border.bottom if old_border and old_border.bottom else Side(style='thin')
        )
        print(f"    Правая граница: ячейка {cell.coordinate}")
    
    print(f"  ✓ Применены жирные границы по периметру группы {start_row}-{end_row}")

def apply_thick_borders_to_supplier_columns(worksheet, sheet_names, max_row):
    """
    Применяет жирные вертикальные границы между колонками поставщиков.
    Каждый поставщик занимает 4 колонки, разделяем их жирными границами.
    
    Args:
        worksheet: Лист Excel для форматирования
        sheet_names: Список имен поставщиков
        max_row: Максимальное количество строк для применения границ
    """
    print(f"Применяем жирные границы для колонок поставщиков. Поставщиков: {len(sheet_names)}")
    
    thick_side = Side(style='thick', color='000000')
    
    # Начинаем с колонки 3 (после "Наименование" и "Количество запрошенное")
    current_col = 3
    
    for i, sheet_name in enumerate(sheet_names):
        # Каждый поставщик занимает 4 колонки
        supplier_start_col = current_col
        supplier_end_col = current_col + 3
        
        print(f"  Поставщик '{sheet_name}': колонки {supplier_start_col}-{supplier_end_col}")
        
        # Применяем жирную левую границу к первой колонке поставщика
        for row in range(1, max_row + 1):
            cell = worksheet.cell(row=row, column=supplier_start_col)
            old_border = cell.border
            cell.border = Border(
                left=thick_side,  # Жирная левая граница
                right=old_border.right if old_border and old_border.right else Side(style='thin'),
                top=old_border.top if old_border and old_border.top else Side(style='thin'),
                bottom=old_border.bottom if old_border and old_border.bottom else Side(style='thin')
            )
        
        # Если это последний поставщик, применяем жирную правую границу к последней колонке
        if i == len(sheet_names) - 1:
            for row in range(1, max_row + 1):
                cell = worksheet.cell(row=row, column=supplier_end_col)
                old_border = cell.border
                cell.border = Border(
                    left=old_border.left if old_border and old_border.left else Side(style='thin'),
                    right=thick_side,  # Жирная правая граница
                    top=old_border.top if old_border and old_border.top else Side(style='thin'),
                    bottom=old_border.bottom if old_border and old_border.bottom else Side(style='thin')
                )
        
        # Переходим к следующему поставщику
        current_col += 4
        
        print(f"    ✓ Применены жирные границы для поставщика '{sheet_name}'")
    
    print(f"✓ Жирные границы для колонок поставщиков применены")

def is_yellow_cell(cell):
    """
    Проверяет, имеет ли ячейка желтую заливку.
    
    Args:
        cell: Ячейка Excel для проверки
        
    Returns:
        bool: True если ячейка имеет желтую заливку
    """
    if cell.fill and cell.fill.start_color:
        # Проверяем различные варианты желтого цвета
        color = cell.fill.start_color.index
        if isinstance(color, str):
            # Желтые цвета в различных форматах
            yellow_colors = ['FFFFFF00', 'FFFF00', 'FFFF99', 'FFFFCC', 'FFC000', 'FFFF66']
            return color.upper() in yellow_colors
    return False

def get_item_type(cell, main_product_name):
    """
    Определяет тип товара: основной, вариант или аналог.
    
    Args:
        cell: Ячейка Excel для проверки
        main_product_name: Название основного товара для сравнения
        
    Returns:
        str: 'main', 'variant' или 'analog'
    """
    # Если ячейка не имеет желтой заливки и отступов - это основной товар
    if not (is_yellow_cell(cell) or (isinstance(cell.value, str) and cell.value.startswith('      '))):
        return 'main'
    
    # Если есть желтая заливка или отступы - это вариант или аналог
    if is_yellow_cell(cell) or (isinstance(cell.value, str) and cell.value.startswith('      ')):
        clean_name = cell.value.strip() if isinstance(cell.value, str) else str(cell.value)
        if clean_name == main_product_name:
            return 'variant'
        else:
            return 'analog'
    
    return 'main'

def is_analog_cell(cell):
    """
    Проверяет, является ли ячейка аналогом или вариантом (желтая заливка ИЛИ начинается с отступов).
    
    Args:
        cell: Ячейка Excel для проверки
        
    Returns:
        bool: True если ячейка является аналогом или вариантом
    """
    # Проверяем желтую заливку
    if is_yellow_cell(cell):
        return True
    
    # Проверяем отступы (начинается с пробелов)
    if isinstance(cell.value, str) and cell.value.startswith('      '):
        return True
    
    return False

def find_best_main_product_for_analog(analog_name, main_products_list, all_data_sequence, analog_qty=None):
    """
    Находит наиболее подходящий основной товар для аналога на основе текстового сопоставления и количества.
    
    Args:
        analog_name: Название аналога
        main_products_list: Список названий основных товаров
        all_data_sequence: Все данные для анализа позиций
        analog_qty: Количество аналога (опционально)
        
    Returns:
        str: Название наиболее подходящего основного товара
    """
    if not main_products_list:
        return None
    
    # Создаем словарь количеств основных товаров
    main_product_quantities = {}
    for item in all_data_sequence:
        item_type = get_item_type(item['name_cell'], "")
        if item_type == 'main':
            main_product_quantities[item['product_name']] = item['requested_qty']
    
    best_product = None
    best_similarity = 0
    
    print(f"\n--- ПОИСК ЛУЧШЕГО ОСНОВНОГО ТОВАРА ДЛЯ АНАЛОГА: '{analog_name}' (кол-во: {analog_qty}) ---")
    
    # Проверяем каждый основной товар
    for main_product in main_products_list:
        main_qty = main_product_quantities.get(main_product, None)
        
        # Рассчитываем сходство с учетом количества
        similarity = calculate_weighted_similarity(analog_name, main_product, qty1=analog_qty, qty2=main_qty)
        print(f"  Сходство с '{main_product}' (кол-во: {main_qty}): {similarity:.3f}")
        
        # Проверяем, можно ли группировать с учетом количества
        can_group = should_group_items(similarity, analog_qty, main_qty)
        
        if can_group and similarity > best_similarity:
            best_similarity = similarity
            best_product = main_product
            print(f"    → НОВЫЙ ЛУЧШИЙ: {main_product} (сходство: {similarity:.3f})")
        elif not can_group:
            print(f"    → ОТКЛОНЕН по количеству")
    
    if best_product:
        print(f"  ИТОГ: выбран '{best_product}' с сходством {best_similarity:.3f}")
        return best_product
    else:
        print(f"  ИТОГ: подходящий основной товар НЕ найден")
        return None

def extract_payment_terms(wb, sheet_names):
    """
    Извлекает условия оплаты с первого листа для каждого поставщика.
    Ищет строку "условия оплаты" и извлекает условия из объединенных ячеек справа.
    
    Args:
        wb: Рабочая книга Excel
        sheet_names: Список имен листов поставщиков
        
    Returns:
        dict: Словарь {sheet_name: payment_terms}
    """
    payment_terms = {}
    
    # Получаем первый лист (обычно это лист с общей информацией)
    first_sheet_name = wb.sheetnames[0]
    first_ws = wb[first_sheet_name]
    
    # Ищем строку "условия оплаты"
    payment_row = None
    payment_col = None
    
    for row in range(1, first_ws.max_row + 1):
        for col in range(1, first_ws.max_column + 1):
            cell = first_ws.cell(row, col)
            if cell.value and isinstance(cell.value, str):
                cell_value = cell.value.strip().lower()
                if 'условия оплаты' in cell_value:
                    payment_row = row
                    payment_col = col
                    print(f"Найдена строка 'условия оплаты' в ячейке {cell.coordinate}")
                    break
        if payment_row:
            break
    
    if not payment_row:
        print("Строка 'условия оплаты' не найдена на первом листе")
        return payment_terms
    
    # После нахождения строки "условия оплаты", ищем условия для каждого поставщика
    # в объединенных ячейках справа от найденной строки
    
    # Сначала найдем заголовки поставщиков, чтобы понять их расположение
    supplier_columns = {}  # {sheet_name: column_range}
    
    # Ищем заголовки поставщиков в первых строках
    for row in range(1, min(10, first_ws.max_row + 1)):
        for merged_range in first_ws.merged_cells.ranges:
            if merged_range.min_row == row:
                top_left_cell = first_ws.cell(merged_range.min_row, merged_range.min_col)
                if top_left_cell.value and isinstance(top_left_cell.value, str):
                    header_value = top_left_cell.value.strip()
                    
                    # Проверяем, соответствует ли заголовок одному из поставщиков
                    for sheet_name in sheet_names:
                        sheet_name_clean = sheet_name.replace('"', '').replace("'", '').strip()
                        if (sheet_name_clean.lower() in header_value.lower() or
                            header_value.lower() in sheet_name_clean.lower() or
                            any(word in header_value.lower() for word in sheet_name_clean.lower().split() if len(word) > 2)):
                            supplier_columns[sheet_name] = {
                                'start_col': merged_range.min_col,
                                'end_col': merged_range.max_col,
                                'header': header_value
                            }
                            print(f"Найден поставщик '{sheet_name}' в колонках {merged_range.min_col}-{merged_range.max_col}")
                            break
    
    # Теперь ищем условия оплаты в строке payment_row для каждого поставщика
    for sheet_name, col_info in supplier_columns.items():
        found_payment_terms = False
        
        # Ищем объединенные ячейки в строке условий оплаты для данного поставщика
        for merged_range in first_ws.merged_cells.ranges:
            # Проверяем, пересекается ли объединенная ячейка с колонками поставщика
            # и находится ли она в строке условий оплаты или рядом с ней
            if (merged_range.min_row >= payment_row and merged_range.min_row <= payment_row + 3 and
                merged_range.min_col >= col_info['start_col'] and
                merged_range.min_col <= col_info['end_col']):  # Строгая проверка пересечения
                
                top_left_cell = first_ws.cell(merged_range.min_row, merged_range.min_col)
                if top_left_cell.value and isinstance(top_left_cell.value, str):
                    cell_value = top_left_cell.value.strip()
                    # Исключаем саму строку "условия оплаты"
                    if 'условия оплаты' not in cell_value.lower() and len(cell_value) > 3:
                        payment_terms[sheet_name] = cell_value
                        print(f"Найдены условия оплаты для '{sheet_name}' в объединенной ячейке {merged_range}: {cell_value}")
                        found_payment_terms = True
                        break
        
        # Если не нашли в объединенных ячейках, ищем в обычных ячейках
        if not found_payment_terms:
            for check_row in range(payment_row, payment_row + 4):  # Проверяем несколько строк после "условия оплаты"
                for check_col in range(col_info['start_col'], col_info['end_col'] + 1):
                    if check_row <= first_ws.max_row and check_col <= first_ws.max_column:
                        cell = first_ws.cell(check_row, check_col)
                        if cell.value and isinstance(cell.value, str):
                            cell_value = cell.value.strip()
                            if ('условия оплаты' not in cell_value.lower() and
                                len(cell_value) > 3 and len(cell_value) < 200):
                                payment_terms[sheet_name] = cell_value
                                print(f"Найдены условия оплаты для '{sheet_name}' в обычной ячейке ({check_row}, {check_col}): {cell_value}")
                                found_payment_terms = True
                                break
                if found_payment_terms:
                    break
        
        # Если условия оплаты не найдены для поставщика, выводим сообщение
        if not found_payment_terms:
            print(f"Условия оплаты для '{sheet_name}' НЕ найдены в колонках {col_info['start_col']}-{col_info['end_col']}")
    
    return payment_terms
    
def clean_text_for_comparison(text):
    """
    Очищает текст для сравнения: убирает лишние символы, приводит к нижнему регистру.
    
    Args:
        text: Исходный текст
        
    Returns:
        str: Очищенный текст
    """
    import re
    if not isinstance(text, str):
        text = str(text)
    
    # Приводим к нижнему регистру
    text = text.lower()
    
    # Убираем лишние символы, оставляем только буквы, цифры и пробелы
    text = re.sub(r'[^\w\s]', ' ', text)
    
    # Убираем множественные пробелы
    text = re.sub(r'\s+', ' ', text).strip()
    
    return text
    
def determine_word_weight(word):
    """
    Определяет вес слова для сравнения товаров.
    
    Args:
        word: Слово для анализа
        
    Returns:
        int: Вес слова (1, 2 или 3)
    """
    import re
    
    # Ключевые слова категорий (вес x3)
    category_keywords = {
        'зарядное', 'зарядка', 'адаптер', 'блок', 'питания',
        'кабель', 'провод', 'шнур', 'cord', 'cable',
        'накопитель', 'диск', 'ssd', 'hdd', 'память', 'storage', 'твердотельный',
        'телевизор', 'тв', 'tv', 'oled', 'led', 'qled',
        'консоль', 'playstation', 'xbox', 'геймпад', 'джойстик',
        'наушники', 'гарнитура', 'headphones', 'earphones',
        'мышь', 'клавиатура', 'mouse', 'keyboard'
    }
    
    # Технические характеристики (вес x2)
    technical_patterns = [
        r'\d+gb', r'\d+tb', r'\d+мб', r'\d+гб',  # Объем памяти
        r'\d+"', r'\d+дюйм',  # Размеры экранов
        r'\d+вт', r'\d+w',  # Мощность
        r'usb', r'type-c', r'lightning', r'hdmi',  # Интерфейсы
        r'\d+hz', r'\d+гц',  # Частота
        r'4k', r'8k', r'hd', r'fullhd',  # Разрешение
        r'\d+a', r'\d+ампер',  # Ток
        r'\d+v', r'\d+вольт'  # Напряжение
    ]
    
    word_lower = word.lower()
    
    # Проверяем ключевые слова категорий
    if word_lower in category_keywords:
        return 3
    
    # Проверяем технические характеристики
    for pattern in technical_patterns:
        if re.search(pattern, word_lower):
            return 2
    
    # Обычные слова
    return 1
    
def calculate_weighted_similarity(text1, text2, qty1=None, qty2=None):
    """
    Рассчитывает взвешенное сходство между двумя текстами с учетом синонимов и количества.
    ПРИОРИТЕТ: Сначала сравниваются количества, затем текстовое сходство.
    
    Args:
        text1: Первый текст для сравнения (обычно аналог)
        text2: Второй текст для сравнения (обычно основной товар)
        qty1: Запрашиваемое количество для первого товара (опционально)
        qty2: Запрашиваемое количество для второго товара (опционально)
        
    Returns:
        float: Коэффициент сходства от 0.0 до 1.0
    """
    # ПРИОРИТЕТНАЯ ПРОВЕРКА КОЛИЧЕСТВА
    qty_similarity = 0.0
    has_quantity_data = False
    
    if qty1 is not None and qty2 is not None:
        try:
            # Приводим к числам для сравнения
            q1 = float(qty1) if qty1 != '' else 0
            q2 = float(qty2) if qty2 != '' else 0
            has_quantity_data = True
            
            if q1 == q2 and q1 > 0:
                # ТОЧНОЕ совпадение количества - максимальный приоритет
                qty_similarity = 1.0
                print(f"  ⭐ ТОЧНОЕ совпадение количества: {q1} = {q2} (приоритет: 1.0)")
            elif q1 != q2 and q1 > 0 and q2 > 0:
                # Количества не совпадают - сильно снижаем приоритет
                qty_similarity = 0.1  # Очень низкий приоритет для разных количеств
                print(f"  ❌ Количества НЕ совпадают: {q1} ≠ {q2} (приоритет: 0.1)")
            else:
                # Одно из количеств равно 0 или пустое
                qty_similarity = 0.3  # Средний приоритет
                print(f"  ⚠️ Неполные данные о количестве: {q1}, {q2} (приоритет: 0.3)")
        except (ValueError, TypeError):
            # Если не удалось преобразовать в числа
            qty_similarity = 0.3
            has_quantity_data = False
            print(f"  ⚠️ Ошибка обработки количества: {qty1}, {qty2} (приоритет: 0.3)")
    else:
        # Нет данных о количестве
        qty_similarity = 0.5  # Нейтральный приоритет
        print(f"  ℹ️ Нет данных о количестве (приоритет: 0.5)")
    # Словарь синонимов для лучшего сопоставления
    synonyms = {
        'ssd': ['накопитель', 'твердотельный', 'диск'],
        'накопитель': ['ssd', 'твердотельный', 'диск'],
        'твердотельный': ['ssd', 'накопитель', 'диск'],
        'диск': ['ssd', 'накопитель', 'твердотельный'],
        'телевизор': ['тв', 'tv'],
        'тв': ['телевизор', 'tv'],
        'tv': ['телевизор', 'тв'],
        'кабель': ['провод', 'шнур'],
        'провод': ['кабель', 'шнур'],
        'шнур': ['кабель', 'провод'],
        'зарядное': ['зарядка', 'адаптер'],
        'зарядка': ['зарядное', 'адаптер'],
        'адаптер': ['зарядное', 'зарядка'],
        'устройство': ['девайс', 'прибор'],
        'девайс': ['устройство', 'прибор'],
        'прибор': ['устройство', 'девайс'],
        # Добавляем синонимы для АКБ и батарей
        'акб': ['батарея', 'аккумулятор', 'аккумуляторная'],
        'батарея': ['акб', 'аккумулятор', 'аккумуляторная'],
        'аккумулятор': ['акб', 'батарея', 'аккумуляторная'],
        'аккумуляторная': ['акб', 'батарея', 'аккумулятор'],
        # Добавляем синонимы для ТСД
        'тсд': ['терминал', 'сбора', 'данных'],
        'терминал': ['тсд'],
        # Добавляем синонимы для ЗУ и зарядных устройств
        'зу': ['зарядное', 'устройство', 'зарядка', 'кредл', 'зарядный'],
        'зарядное': ['зу', 'зарядка', 'адаптер', 'кредл', 'зарядный'],
        'устройство': ['зу', 'девайс', 'прибор'],
        'кредл': ['зу', 'зарядное', 'зарядка', 'зарядный'],
        'зарядный': ['зу', 'зарядное', 'кредл', 'зарядка']
    }
    
    # Очищаем тексты
    clean_text1 = clean_text_for_comparison(text1)
    clean_text2 = clean_text_for_comparison(text2)
    
    # Разбиваем на слова
    words1 = set(clean_text1.split())
    words2 = set(clean_text2.split())
    
    # Исключаем служебные слова
    stop_words = {'и', 'или', 'с', 'для', 'на', 'в', 'от', 'до', 'по', 'без', 'при', 'под', 'над', 'за', 'к', 'у'}
    words1 = words1 - stop_words
    words2 = words2 - stop_words
    
    # Исключаем слишком короткие слова (менее 3 символов)
    words1 = {w for w in words1 if len(w) >= 3}
    words2 = {w for w in words2 if len(w) >= 3}
    
    if not words1 or not words2:
        return 0.0
    
    # Рассчитываем взвешенные баллы
    total_weight1 = sum(determine_word_weight(word) for word in words1)
    total_weight2 = sum(determine_word_weight(word) for word in words2)
    
    # Находим общие слова с учетом синонимов
    common_words = words1.intersection(words2)
    
    # Добавляем синонимичные пары
    synonym_matches = set()
    for word1 in words1:
        if word1 in synonyms:
            for synonym in synonyms[word1]:
                if synonym in words2:
                    synonym_matches.add(word1)
                    synonym_matches.add(synonym)
                    print(f"  Найдена синонимичная пара: '{word1}' ↔ '{synonym}'")
    
    # Объединяем прямые совпадения и синонимичные пары
    all_common_words = common_words.union(synonym_matches)
    common_weight = sum(determine_word_weight(word) for word in all_common_words)
    
    if total_weight1 == 0 or total_weight2 == 0:
        return 0.0
    
    # Рассчитываем взвешенное сходство (среднее от двух направлений)
    similarity1 = common_weight / total_weight1
    similarity2 = common_weight / total_weight2
    
    text_similarity = (similarity1 + similarity2) / 2
    
    # КОМБИНИРОВАННАЯ ОЦЕНКА: Количество имеет приоритет над текстом
    if has_quantity_data and qty_similarity == 1.0:
        # Точное совпадение количества - текстовое сходство становится вторичным
        final_similarity = 0.7 + (text_similarity * 0.3)  # 70% за количество + 30% за текст
        print(f"  🎯 ПРИОРИТЕТ по количеству: итоговое сходство = {final_similarity:.3f}")
    elif has_quantity_data and qty_similarity == 0.1:
        # Разные количества - сильно снижаем итоговую оценку
        final_similarity = text_similarity * 0.2  # Только 20% от текстового сходства
        print(f"  ⬇️ ШТРАФ за разные количества: итоговое сходство = {final_similarity:.3f}")
    else:
        # Стандартная логика: комбинируем количество и текст
        final_similarity = (qty_similarity * 0.4) + (text_similarity * 0.6)  # 40% количество + 60% текст
        print(f"  ⚖️ СТАНДАРТНАЯ оценка: итоговое сходство = {final_similarity:.3f}")
    
    # Ограничиваем результат до 1.0
    final_similarity = min(1.0, final_similarity)
    
    return final_similarity

def should_group_items(similarity, qty1, qty2):
    """
    Определяет, следует ли группировать товары на основе сходства и количества.
    Применяет разные пороги для товаров с одинаковыми и разными количествами.
    
    Args:
        similarity: Коэффициент сходства от 0.0 до 1.0
        qty1: Количество первого товара
        qty2: Количество второго товара
        
    Returns:
        bool: True если товары следует группировать
    """
    base_threshold = 0.25  # Базовый порог для одинаковых количеств
    different_qty_threshold = 0.7  # Высокий порог для разных количеств
    
    # Проверяем, известны ли количества и различаются ли они
    if qty1 is not None and qty2 is not None:
        try:
            q1 = float(qty1) if qty1 != '' else 0
            q2 = float(qty2) if qty2 != '' else 0
            
            # Если количества известны, больше 0 и не равны - применяем высокий порог
            if q1 > 0 and q2 > 0 and q1 != q2:
                threshold = different_qty_threshold
                print(f"  📊 РАЗНЫЕ количества ({q1} ≠ {q2}) → порог {threshold:.0%}, сходство {similarity:.3f}")
                return similarity >= threshold
            else:
                threshold = base_threshold
                print(f"  📊 Одинаковые/неизвестные количества → порог {threshold:.0%}, сходство {similarity:.3f}")
                return similarity >= threshold
        except (ValueError, TypeError):
            # Если не удалось преобразовать в числа, используем базовый порог
            print(f"  📊 Ошибка обработки количества → базовый порог {base_threshold:.0%}, сходство {similarity:.3f}")
            return similarity >= base_threshold
    else:
        # Нет данных о количестве - используем базовый порог
        print(f"  📊 Нет данных о количестве → базовый порог {base_threshold:.0%}, сходство {similarity:.3f}")
        return similarity >= base_threshold
    
def generate_main_product_name(analog_name):
    """
    Генерирует стандартное название основного товара на основе категории аналога.
    
    Args:
        analog_name: Название аналога
        
    Returns:
        str: Стандартное название основного товара для данной категории
    """
    # Очищаем название
    clean_name = clean_text_for_comparison(analog_name)
    words = set(clean_name.split())
    
    # Определяем категории по ключевым словам (порядок важен - более специфичные категории первыми)
    categories = {
        'монитор': {
            'keywords': ['монитор', 'monitor', 'dell', 'acer', 'samsung', 'philips', 'bravus', 'lime', 'ips', 'lcd', 'led', 'p2422he', 'p2423de', 'u2424h', 'b247y', 's24a604', 'ut241y', '24b2n4200', 'bvq2737pc', 't238a'],
            'standard_name': 'Монитор'
        },
        'телевизор': {
            'keywords': ['телевизор', 'тв', 'tv'],  # Убрал oled, led, qled - они есть и в мониторах
            'standard_name': 'Телевизор OLED'
        },
        'кабель': {
            'keywords': ['кабель', 'провод', 'шнур', 'cable', 'cord'],
            'standard_name': 'Кабель USB'
        },
        'накопитель': {
            'keywords': ['накопитель', 'ssd', 'hdd', 'твердотельный', 'диск', 'storage'],
            'standard_name': 'SSD Накопитель'
        },
        'зарядка': {
            'keywords': ['зарядное', 'зарядка', 'адаптер', 'блок', 'питания'],
            'standard_name': 'Зарядное Устройство'
        },
        'консоль': {
            'keywords': ['playstation', 'xbox', 'консоль', 'ps5', 'геймпад'],
            'standard_name': 'Игровая Консоль'
        },
        'наушники': {
            'keywords': ['наушники', 'гарнитура', 'headphones', 'earphones'],
            'standard_name': 'Наушники'
        }
    }
    
    # Ищем подходящую категорию
    for category_name, category_info in categories.items():
        keywords = category_info['keywords']
        # Проверяем, есть ли ключевые слова этой категории в названии товара
        # Используем более гибкий поиск - проверяем вхождение подстроки
        found_keywords = []
        for keyword in keywords:
            if any(keyword.lower() in word.lower() for word in clean_name.split()):
                found_keywords.append(keyword)
        
        if found_keywords:
            print(f"  Определена категория '{category_name}' для аналога '{analog_name[:50]}...' по ключевым словам: {found_keywords}")
            return category_info['standard_name']
    
    # Если категория не определена, используем старый алгоритм
    print(f"  Категория не определена для аналога '{analog_name[:50]}...', используем старый алгоритм")
    
    # Оставляем только важные слова (вес >= 2)
    important_words = []
    for word in clean_name.split():
        if determine_word_weight(word) >= 2:
            important_words.append(word)
    
    # Если важных слов мало, добавляем обычные слова
    if len(important_words) < 2:
        for word in clean_name.split():
            if determine_word_weight(word) == 1 and len(word) >= 4:
                important_words.append(word)
                if len(important_words) >= 3:
                    break
    
    # Формируем название основного товара
    if important_words:
        result = ' '.join(important_words[:3])  # Берем максимум 3 слова
        return result.title()  # Приводим к красивому виду
    else:
        # Если не удалось выделить важные слова, берем первые 3 слова
        return ' '.join(clean_name.split()[:3]).title()
    
def find_misplaced_analogs(main_product_name, analogs_list, main_product_qty=None):
    """
    Находит аналоги, которые не подходят к основному товару.
    Использует универсальную функцию should_group_items для проверки.
    
    Args:
        main_product_name: Название основного товара
        analogs_list: Список аналогов для проверки
        main_product_qty: Количество основного товара
        
    Returns:
        list: Список неподходящих аналогов
    """
    misplaced_analogs = []
    
    for analog in analogs_list:
        # Получаем количество аналога (если доступно)
        analog_qty = analog.get('requested_qty', None)
        similarity = calculate_weighted_similarity(main_product_name, analog['name'], qty1=main_product_qty, qty2=analog_qty)
        
        # Используем универсальную функцию для проверки
        if not should_group_items(similarity, main_product_qty, analog_qty):
            misplaced_analogs.append({
                'analog': analog,
                'similarity': similarity,
                'suggested_main_product': generate_main_product_name(analog['name']),
                'reason': f'Не прошел проверку группировки (сходство: {similarity:.1%})'
            })
            print(f"Найден неподходящий аналог: '{analog['name']}' для основного товара '{main_product_name}' (сходство: {similarity:.1%})")
    
    return misplaced_analogs

def collect_data_sequentially(wb, sheet_names):
    """
    Последовательно собирает данные товар за товаром в правильном порядке.
    Правильно обрабатывает основные товары, их варианты и аналоги.
    
    Args:
        wb: Рабочая книга Excel
        sheet_names: Список имен листов для обработки
        
    Returns:
        list: Список строк для сводной таблицы в правильном порядке
    """
    summary_rows = []
    
    # ЭТАП 1: Собираем все данные по листам в правильном порядке
    all_data_sequence = []
    
    # Проходим по всем листам и собираем данные в том порядке, как они идут
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if not row[0].value:
                continue
            
            name_cell = row[0]
            product_name = name_cell.value.strip() if isinstance(name_cell.value, str) else str(name_cell.value)
            row_data = [cell.value for cell in row]
            
            all_data_sequence.append({
                'sheet_name': sheet_name,
                'row_idx': row_idx,
                'name_cell': name_cell,
                'product_name': product_name,
                'requested_qty': row_data[1],
                'offered_data': row_data[2:6]
            })
    
    # ЭТАП 2: Находим все основные товары в порядке их первого появления
    main_products_order = []
    seen_main_products = set()
    
    for item in all_data_sequence:
        # Для определения типа используем пустую строку как основной товар
        item_type = get_item_type(item['name_cell'], "")
        if item_type == 'main' and item['product_name'] not in seen_main_products:
            main_products_order.append(item['product_name'])
            seen_main_products.add(item['product_name'])
    
    # ЭТАП 2.1: Находим "сиротские" аналоги (аналоги без основного товара)
    # Это могут быть аналоги ТВ или любых других товаров
    orphan_analogs = []
    tv_main_products = []  # Инициализируем переменную
    
    # Проходим по всем данным и ищем аналоги, которые не имеют основного товара
    for item in all_data_sequence:
        item_type = get_item_type(item['name_cell'], "")
        if item_type == 'analog':
            # Проверяем, есть ли основной товар для этого аналога
            has_main_product = False
            
            # Ищем в том же листе основной товар перед этим аналогом
            sheet_items = [x for x in all_data_sequence if x['sheet_name'] == item['sheet_name']]
            for sheet_item in sheet_items:
                if sheet_item['row_idx'] >= item['row_idx']:
                    break
                sheet_item_type = get_item_type(sheet_item['name_cell'], "")
                if sheet_item_type == 'main':
                    has_main_product = True
                    break
            
            # Если нет основного товара, это "сиротский" аналог
            if not has_main_product:
                analog_info = {
                    'name': item['product_name'],
                    'sheet_name': item['sheet_name'],
                    'row_idx': item['row_idx']
                }
                if analog_info not in orphan_analogs:
                    orphan_analogs.append(analog_info)
    
    # Группируем сиротские аналоги и создаем для них основные товары
    orphan_analogs_by_main_product = {}  # Группируем аналоги по основным товарам
    tv_main_products = []
    
    
    if orphan_analogs:
        # Разделяем аналоги на ТВ и обычные товары
        tv_analogs = [a for a in orphan_analogs if any(tv_keyword in a['name'].upper() for tv_keyword in ['OLED', 'LG', 'ТЕЛЕВИЗОР'])]
        regular_analogs = [a for a in orphan_analogs if not any(tv_keyword in a['name'].upper() for tv_keyword in ['OLED', 'LG', 'ТЕЛЕВИЗОР'])]
        
        # Обрабатываем обычные сиротские аналоги
        if regular_analogs and main_products_order:
            for analog in regular_analogs:
                # Получаем количество аналога
                analog_qty = None
                for item in all_data_sequence:
                    if item['product_name'] == analog['name']:
                        analog_qty = item['requested_qty']
                        break
                
                # Находим наиболее подходящий основной товар для этого аналога
                best_main_product = find_best_main_product_for_analog(
                    analog['name'],
                    main_products_order,
                    all_data_sequence,
                    analog_qty
                )
                
                if best_main_product:
                    if best_main_product not in orphan_analogs_by_main_product:
                        orphan_analogs_by_main_product[best_main_product] = []
                    orphan_analogs_by_main_product[best_main_product].append(analog)
    
    # ЭТАП 2.2: Собираем ВСЕ аналоги для текстового сопоставления (исключаем варианты)
    all_analogs_for_matching = {}  # {analog_name: [analog_data, ...]}
    
    for item in all_data_sequence:
        item_type = get_item_type(item['name_cell'], "")
        if item_type == 'analog':
            # Проверяем, не является ли этот "аналог" на самом деле вариантом
            # (т.е. его название совпадает с каким-то основным товаром)
            is_variant = False
            for main_product in main_products_order:
                if item['product_name'] == main_product:
                    is_variant = True
                    break
            
            # Добавляем только настоящие аналоги (не варианты)
            if not is_variant:
                analog_name = item['product_name']
                if analog_name not in all_analogs_for_matching:
                    all_analogs_for_matching[analog_name] = []
                all_analogs_for_matching[analog_name].append(item)
    
    # ЭТАП 2.3: Умное сопоставление аналогов с проверкой сходства
    print("\n=== УМНОЕ СОПОСТАВЛЕНИЕ АНАЛОГОВ ===")
    
    analogs_by_main_product = {}
    virtual_main_products = {}  # Для хранения виртуальных основных товаров
    
    # Обрабатываем аналоги по одному, чтобы учитывать уже созданные виртуальные товары
    for analog_name, analog_items in all_analogs_for_matching.items():
        # Проверяем сходство со ВСЕМИ основными товарами (включая уже созданные виртуальные)
        best_similarity = 0
        best_main_product = None
        is_virtual = False
        
        # Получаем количество аналога
        analog_qty = analog_items[0]['requested_qty'] if analog_items else None
        
        # Создаем словарь количеств основных товаров для быстрого поиска
        main_product_quantities = {}
        for item in all_data_sequence:
            item_type = get_item_type(item['name_cell'], "")
            if item_type == 'main':
                main_product_quantities[item['product_name']] = item['requested_qty']
        
        print(f"\n--- АНАЛИЗ АНАЛОГА: '{analog_name}' (кол-во: {analog_qty}) ---")
        
        # Проверяем сходство с исходными основными товарами
        for main_product in main_products_order:
            # Находим количество основного товара
            main_qty = main_product_quantities.get(main_product, None)
            similarity = calculate_weighted_similarity(analog_name, main_product, qty1=analog_qty, qty2=main_qty)
            print(f"  Сходство с '{main_product}' (кол-во: {main_qty}): {similarity:.3f}")
            if similarity > best_similarity:
                best_similarity = similarity
                best_main_product = main_product
                is_virtual = main_product in virtual_main_products
        
        # Также проверяем сходство с уже созданными виртуальными товарами
        for virtual_main_name in virtual_main_products.keys():
            if virtual_main_name not in main_products_order:
                continue  # Уже проверили выше
            # Для виртуальных товаров используем количество первого аналога
            virtual_qty = virtual_main_products[virtual_main_name][0]['items'][0]['requested_qty'] if virtual_main_products[virtual_main_name] else None
            similarity = calculate_weighted_similarity(analog_name, virtual_main_name, qty1=analog_qty, qty2=virtual_qty)
            print(f"  Сходство с виртуальным '{virtual_main_name}' (кол-во: {virtual_qty}): {similarity:.3f}")
            if similarity > best_similarity:
                best_similarity = similarity
                best_main_product = virtual_main_name
                is_virtual = True
        
        print(f"  ИТОГ: лучшее сходство {best_similarity:.3f} с '{best_main_product}' {'(виртуальный)' if is_virtual else '(исходный)'}")
        
        # Решаем: привязать к существующему или создать виртуальный товар
        # Находим количество лучшего основного товара
        best_main_qty = None
        if best_main_product in main_product_quantities:
            best_main_qty = main_product_quantities[best_main_product]
        elif best_main_product in virtual_main_products:
            # Для виртуальных товаров берем количество первого аналога
            if virtual_main_products[best_main_product]:
                best_main_qty = virtual_main_products[best_main_product][0]['items'][0]['requested_qty']
        
        # Используем универсальную функцию для проверки группировки
        if should_group_items(best_similarity, analog_qty, best_main_qty):
            if is_virtual:
                # Привязываем к виртуальному основному товару
                virtual_main_products[best_main_product].append({
                    'name': analog_name,
                    'items': analog_items
                })
                print(f"  → Привязан к виртуальному товару '{best_main_product}'")
            else:
                # Привязываем к исходному основному товару
                if best_main_product not in analogs_by_main_product:
                    analogs_by_main_product[best_main_product] = []
                analogs_by_main_product[best_main_product].append({
                    'name': analog_name,
                    'items': analog_items
                })
                print(f"  → Привязан к исходному товару '{best_main_product}'")
        else:
            # Создаем новый виртуальный основной товар
            virtual_main_name = generate_main_product_name(analog_name)
            
            # Проверяем, не существует ли уже виртуальный товар с таким же названием
            if virtual_main_name in virtual_main_products:
                # Добавляем к существующему виртуальному товару
                virtual_main_products[virtual_main_name].append({
                    'name': analog_name,
                    'items': analog_items
                })
                print(f"  → Аналог добавлен к существующему виртуальному товару: '{virtual_main_name}'")
            else:
                # Создаем новый виртуальный товар
                virtual_main_products[virtual_main_name] = []
                virtual_main_products[virtual_main_name].append({
                    'name': analog_name,
                    'items': analog_items
                })
                
                # Добавляем виртуальный основной товар в общий список
                main_products_order.append(virtual_main_name)
                print(f"  → Создан новый виртуальный основной товар: '{virtual_main_name}'")
    
    print(f"Создано виртуальных основных товаров: {len(virtual_main_products)}")
    
    
    # ЭТАП 3: Обрабатываем каждый основной товар последовательно
    for main_product_name in main_products_order:
        
        # 3.1: Собираем предложения по основному товару
        main_product_offers = {}
        main_requested_qty = None
        
        for item in all_data_sequence:
            item_type = get_item_type(item['name_cell'], main_product_name)
            if item_type == 'main' and item['product_name'] == main_product_name:
                main_product_offers[item['sheet_name']] = item['offered_data']
                if main_requested_qty is None:
                    main_requested_qty = item['requested_qty']
        
        
        # Если нет предложений по основному товару, но есть варианты или аналоги,
        # берем количество из первого варианта/аналога
        if main_requested_qty is None:
            # Ищем количество в вариантах
            for item in all_data_sequence:
                item_type = get_item_type(item['name_cell'], main_product_name)
                if item_type == 'variant' and item['requested_qty'] is not None:
                    main_requested_qty = item['requested_qty']
                    break
            
            # Если не нашли в вариантах, ищем в аналогах
            if main_requested_qty is None:
                if main_product_name in analogs_by_main_product:
                    for analog in analogs_by_main_product[main_product_name]:
                        if analog['items'] and analog['items'][0]['requested_qty'] is not None:
                            main_requested_qty = analog['items'][0]['requested_qty']
                            break
        
        # Добавляем строку основного товара (может быть пустой, если нет предложений)
        main_row = {
            'name': main_product_name,
            'requested_qty': main_requested_qty,
            'suppliers': main_product_offers
        }
        summary_rows.append(main_row)
        
        # 3.2: Ищем и добавляем варианты основного товара
        variant_counter = 1
        
        for item in all_data_sequence:
            item_type = get_item_type(item['name_cell'], main_product_name)
            if item_type == 'variant':
                variant_row = {
                    'name': f"{main_product_name} (вариант {variant_counter})",
                    'requested_qty': item['requested_qty'],
                    'suppliers': {item['sheet_name']: item['offered_data']}
                }
                summary_rows.append(variant_row)
                variant_counter += 1
        
        # 3.3: Ищем и добавляем аналоги основного товара
        analog_counter = 1
        processed_analogs = set()
        
        # Добавляем аналоги, найденные через текстовое сопоставление
        if main_product_name in analogs_by_main_product:
            for analog in analogs_by_main_product[main_product_name]:
                if analog['name'] not in processed_analogs:
                    # Собираем все предложения по этому аналогу из всех листов
                    analog_offers = {}
                    analog_requested_qty = None
                    
                    for analog_item in analog['items']:
                        analog_offers[analog_item['sheet_name']] = analog_item['offered_data']
                        if analog_requested_qty is None:
                            analog_requested_qty = analog_item['requested_qty']
                    
                    analog_row = {
                        'name': f"{analog['name']} (аналог {analog_counter})",
                        'requested_qty': analog_requested_qty,
                        'suppliers': analog_offers
                    }
                    summary_rows.append(analog_row)
                    processed_analogs.add(analog['name'])
                    analog_counter += 1
        
        # Добавляем сиротские аналоги, которые не попали в общее сопоставление
        if main_product_name in orphan_analogs_by_main_product:
            for orphan_analog in orphan_analogs_by_main_product[main_product_name]:
                if orphan_analog['name'] not in processed_analogs:
                    # Собираем все предложения по этому аналогу из всех листов
                    analog_offers = {}
                    analog_requested_qty = None
                    
                    for analog_item in all_data_sequence:
                        if analog_item['product_name'] == orphan_analog['name']:
                            analog_offers[analog_item['sheet_name']] = analog_item['offered_data']
                            if analog_requested_qty is None:
                                analog_requested_qty = analog_item['requested_qty']
                    
                    analog_row = {
                        'name': f"{orphan_analog['name']} (аналог {analog_counter})",
                        'requested_qty': analog_requested_qty,
                        'suppliers': analog_offers
                    }
                    summary_rows.append(analog_row)
                    processed_analogs.add(orphan_analog['name'])
                    analog_counter += 1
        
        # Добавляем аналоги из виртуальных основных товаров
        if main_product_name in virtual_main_products:
            for virtual_analog in virtual_main_products[main_product_name]:
                if virtual_analog['name'] not in processed_analogs:
                    # Собираем все предложения по этому аналогу из всех листов
                    analog_offers = {}
                    analog_requested_qty = None
                    
                    for analog_item in virtual_analog['items']:
                        analog_offers[analog_item['sheet_name']] = analog_item['offered_data']
                        if analog_requested_qty is None:
                            analog_requested_qty = analog_item['requested_qty']
                    
                    analog_row = {
                        'name': f"{virtual_analog['name']} (аналог {analog_counter})",
                        'requested_qty': analog_requested_qty,
                        'suppliers': analog_offers
                    }
                    summary_rows.append(analog_row)
                    processed_analogs.add(virtual_analog['name'])
                    analog_counter += 1
    
    return summary_rows

def build_summary_table(filename):
    # Загружаем исходный Excel
    wb = openpyxl.load_workbook(filename)
    
    # Получаем список листов поставщиков (пропускаем первый лист)
    sheet_names = wb.sheetnames[1:]
    
    # ЭТАП 1: Определяем количество основных товаров
    all_main_products = set()
    
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            if not row[0].value:
                continue
            
            name_cell = row[0]
            # Проверяем, является ли это основным товаром (не имеет желтой заливки и отступов)
            if not (is_yellow_cell(name_cell) or (isinstance(name_cell.value, str) and name_cell.value.startswith('      '))):
                product_name = name_cell.value.strip() if isinstance(name_cell.value, str) else str(name_cell.value)
                all_main_products.add(product_name)
    
    print(f"Найдено основных товаров: {len(all_main_products)}")
    print(f"Список основных товаров: {list(all_main_products)}")
    
    # ЭТАП 2: Выбираем формат свода в зависимости от количества основных товаров
    if len(all_main_products) == 1:
        print("Используется упрощенный формат для одного товара")
        return build_single_product_summary(wb, sheet_names)
    else:
        print(f"Используется стандартный формат для {len(all_main_products)} товаров")
        
        # ЭТАП 2.5: Подсчитываем количество заполненных ценой строк для каждого поставщика
        supplier_filled_counts = {}
        
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            filled_count = 0
            
            for row in ws.iter_rows(min_row=2):
                if not row[0].value:
                    continue
                
                # Проверяем, заполнена ли цена (колонка с индексом 3, т.е. 4-я колонка)
                price_value = row[3].value if len(row) > 3 else None
                
                if price_value is not None:
                    try:
                        # Пытаемся преобразовать в число
                        float(price_value)
                        filled_count += 1
                    except (ValueError, TypeError):
                        # Если не число, пропускаем
                        pass
            
            supplier_filled_counts[sheet_name] = filled_count
            print(f"Поставщик '{sheet_name}': {filled_count} заполненных ценой строк")
        
        # ЭТАП 2.6: Сортируем поставщиков по количеству заполненных строк (по убыванию)
        sheet_names = sorted(sheet_names, key=lambda x: supplier_filled_counts[x], reverse=True)
        print(f"\nПорядок поставщиков после сортировки по заполненности:")
        for i, sheet_name in enumerate(sheet_names, 1):
            print(f"  {i}. '{sheet_name}' - {supplier_filled_counts[sheet_name]} строк")
        
        # Продолжаем со стандартной логикой
        pass

    # Создаём новый файл для свода
    summary_wb = openpyxl.Workbook()
    summary_ws = summary_wb.active
    summary_ws.title = "Свод"

    # Извлекаем условия оплаты с первого листа
    payment_terms = extract_payment_terms(wb, sheet_names)
    
    # Формируем заголовки
    headers_row_1 = ["Наименование", "Количество запрошенное"]
    headers_row_2 = ["Наименование", "Количество запрошенное"]

    col = 3
    for sheet_name in sheet_names:
        headers_row_1.extend([sheet_name, "", "", ""])
        headers_row_2.extend(["Количество предложенное", "Цена без НДС за шт", "Сроки поставки", "Комментарий поставщика"])
        col += 4

    summary_ws.append(headers_row_1)
    summary_ws.append(headers_row_2)

    # Объединение A1:A2 и B1:B2
    summary_ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    summary_ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    from openpyxl.styles import Alignment
    summary_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary_ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

    # Объединение заголовков по поставщикам
    for col in range(3, len(headers_row_1) + 1, 4):
        summary_ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+3)
        summary_ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # ПОСЛЕДОВАТЕЛЬНАЯ ЛОГИКА: Обрабатываем товары один за другим в правильном порядке
    summary_rows = collect_data_sequentially(wb, sheet_names)

    # Заполняем свод
    row_idx = 3
    for row_data in summary_rows:
        summary_ws.cell(row=row_idx, column=1, value=row_data['name'])
        summary_ws.cell(row=row_idx, column=2, value=row_data['requested_qty'])

        col = 3
        for sheet_name in sheet_names:
            if sheet_name in row_data['suppliers']:
                vals = row_data['suppliers'][sheet_name]
                for i in range(4):
                    summary_ws.cell(row=row_idx, column=col+i, value=vals[i])
            col += 4

        row_idx += 1

    # Добавляем строку с суммами
    if row_idx > 3:  # Если есть данные
        # Добавляем пустую строку для разделения
        row_idx += 1
        
        # Добавляем строку "ИТОГО"
        summary_ws.cell(row=row_idx, column=1, value="ИТОГО")
        
        # Вычисляем суммы по колонкам "Цена без НДС за шт"
        col = 3
        for sheet_name in sheet_names:
            # Находим колонку с ценами для текущего поставщика
            price_col = None
            for i, header in enumerate(headers_row_2):
                if header == 'Цена без НДС за шт' and i >= col - 1 and i < col + 3:
                    price_col = i + 1
                    break
            
            if price_col:
                # Создаем формулу суммы для колонки с ценами
                start_row = 3
                end_row = row_idx - 2  # Исключаем пустую строку и строку ИТОГО
                formula = f"=SUM({summary_ws.cell(row=start_row, column=price_col).coordinate}:{summary_ws.cell(row=end_row, column=price_col).coordinate})"
                summary_ws.cell(row=row_idx, column=price_col, value=formula)
            
            col += 4
        
        max_col = len(headers_row_1)
        
        # Применяем валютный формат к столбцам с ценами (включая строку ИТОГО)
        apply_currency_format_to_price_columns(summary_ws, headers_row_2, 3, row_idx)
        
        # Выделяем минимальные цены зеленым цветом (исключая строку ИТОГО)
        highlight_minimum_prices(summary_ws, headers_row_2, 3, row_idx - 2)
        
        # Устанавливаем ширину колонок и включаем перенос текста
        set_column_widths_and_wrap_text(summary_ws, headers_row_2)
        
        # СНАЧАЛА: Применяем базовые тонкие границы ко ВСЕМ ячейкам
        apply_borders_to_range(summary_ws, 1, 1, row_idx, max_col)
        
        # ПОТОМ: Применяем все правила по жирным границам
        # Форматируем заголовки (строки 1-2) - жирные границы и светло-голубая заливка
        format_header_rows(summary_ws, max_col)
        
        # Форматируем группы товаров - жирные границы и увеличенный шрифт для основных товаров
        format_main_product_groups(summary_ws, summary_rows, max_col, start_row=3)
        
        # ИСПРАВЛЕНИЕ: Добавляем жирную нижнюю границу к последней строке данных
        # Последняя строка данных находится перед пустой строкой (row_idx - 2)
        last_data_row = row_idx - 2
        thick_bottom_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thick')
        )
        
        # Применяем жирную нижнюю границу ко всем ячейкам последней строки данных
        for col in range(1, max_col + 1):
            cell = summary_ws.cell(row=last_data_row, column=col)
            old_border = cell.border
            cell.border = Border(
                left=old_border.left if old_border and old_border.left else Side(style='thin'),
                right=old_border.right if old_border and old_border.right else Side(style='thin'),
                top=old_border.top if old_border and old_border.top else Side(style='thin'),
                bottom=Side(style='thick')  # Жирная нижняя граница
            )
        
        print(f"✓ Применена жирная нижняя граница к последней строке данных: {last_data_row}")
        
        # Применяем жирные границы для колонок поставщиков
        apply_thick_borders_to_supplier_columns(summary_ws, sheet_names, row_idx)
        
        # Делаем строку ИТОГО жирной
        from openpyxl.styles import Font
        bold_font = Font(bold=True)
        for col_idx in range(1, max_col + 1):
            summary_ws.cell(row=row_idx, column=col_idx).font = bold_font
        
        # Добавляем условия оплаты
        if payment_terms:
            # Добавляем только одну пустую строку для разделения
            row_idx += 1
            
            # Добавляем заголовок "Условия оплаты" в колонку A
            summary_ws.cell(row=row_idx, column=1, value="Условия оплаты:")
            summary_ws.cell(row=row_idx, column=1).font = bold_font
            
            # НЕ увеличиваем row_idx - условия поставщиков должны быть в той же строке!
            
            # Добавляем условия оплаты для каждого поставщика в соответствующих колонках ТОЙ ЖЕ СТРОКИ
            col = 3  # Начинаем с колонки C (первый поставщик)
            for sheet_name in sheet_names:
                if sheet_name in payment_terms:
                    # Каждый поставщик занимает 4 колонки (количество, цена, сроки, комментарий)
                    start_col = col
                    end_col = col + 3
                    
                    # Объединяем ячейки для условий оплаты поставщика (4 колонки)
                    summary_ws.merge_cells(start_row=row_idx, start_column=start_col,
                                         end_row=row_idx, end_column=end_col)
                    
                    # Добавляем текст условий оплаты
                    summary_ws.cell(row=row_idx, column=start_col, value=payment_terms[sheet_name])
                    
                    # Применяем выравнивание и перенос текста
                    summary_ws.cell(row=row_idx, column=start_col).alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    
                    print(f"Добавлены условия оплаты для '{sheet_name}' в колонки {start_col}-{end_col}: {payment_terms[sheet_name]}")
                else:
                    print(f"Условия оплаты для '{sheet_name}' не найдены")
                
                # Переходим к следующему поставщику (следующие 4 колонки)
                col += 4
            
            # Применяем тонкие границы только к новым строкам с условиями оплаты
            apply_borders_to_range(summary_ws, row_idx, 1, row_idx, max_col)
    else:
        max_col = len(headers_row_1)
        
        # Применяем валютный формат к столбцам с ценами
        apply_currency_format_to_price_columns(summary_ws, headers_row_2, 3, row_idx - 1)
        
        # Выделяем минимальные цены зеленым цветом
        highlight_minimum_prices(summary_ws, headers_row_2, 3, row_idx - 1)
        
        # Устанавливаем ширину колонок и включаем перенос текста
        set_column_widths_and_wrap_text(summary_ws, headers_row_2)
        
        # СНАЧАЛА: Применяем базовые тонкие границы ко ВСЕМ ячейкам
        apply_borders_to_range(summary_ws, 1, 1, row_idx - 1, max_col)
        
        # ПОТОМ: Применяем все правила по жирным границам
        # Форматируем заголовки (строки 1-2) - жирные границы и светло-голубая заливка
        format_header_rows(summary_ws, max_col)
        
        # Форматируем группы товаров - жирные границы и увеличенный шрифт для основных товаров
        format_main_product_groups(summary_ws, summary_rows, max_col, start_row=3)
        
        # ИСПРАВЛЕНИЕ: Добавляем жирную нижнюю границу к последней строке данных
        # В этом случае последняя строка данных - это row_idx - 1
        if row_idx > 3:  # Проверяем, что есть данные
            last_data_row = row_idx - 1
            thick_bottom_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thick')
            )
            
            # Применяем жирную нижнюю границу ко всем ячейкам последней строки данных
            for col in range(1, max_col + 1):
                cell = summary_ws.cell(row=last_data_row, column=col)
                old_border = cell.border
                cell.border = Border(
                    left=old_border.left if old_border and old_border.left else Side(style='thin'),
                    right=old_border.right if old_border and old_border.right else Side(style='thin'),
                    top=old_border.top if old_border and old_border.top else Side(style='thin'),
                    bottom=Side(style='thick')  # Жирная нижняя граница
                )
            
            print(f"✓ Применена жирная нижняя граница к последней строке данных: {last_data_row}")
        
        # Применяем жирные границы для колонок поставщиков
        apply_thick_borders_to_supplier_columns(summary_ws, sheet_names, row_idx - 1)

    return summary_wb


def build_single_product_summary(wb, sheet_names):
    """Создает сводную таблицу для случая с одним основным товаром"""
    try:
        summary_wb = openpyxl.Workbook()
        ws = summary_wb.active
        ws.title = "Свод"
    
        # Заголовки
        headers = [
            "Наименование",
            "Количество запрошенное",
            "Количество предложенное",
            "Цена за единицу без НДС в валюте",
            "Цена за единицу без НДС в рублях",
            "Сумма рубли без НДС",
            "Название поставщика",
            "Сроки поставки",
            "Условия оплаты",
            "Комментарий поставщика"
        ]
        ws.append(headers)
        
        # Форматируем заголовки: светло-голубая заливка и жирный шрифт
        light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        bold_font = Font(bold=True)
        
        for col in range(1, 11):  # Колонки A-J (1-10)
            cell = ws.cell(row=1, column=col)
            cell.fill = light_blue_fill
            cell.font = bold_font
        
        # Получаем условия оплаты
        payment_terms = extract_payment_terms(wb, sheet_names)
        
        # Собираем данные по поставщикам
        row_num = 2
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Пропускаем пустые строки
                    continue
                    
                # Добавляем данные поставщика
                ws.cell(row=row_num, column=1, value=row[0])  # Наименование
                ws.cell(row=row_num, column=2, value=row[1])  # Кол-во запрошенное
                ws.cell(row=row_num, column=3, value=row[2])  # Кол-во предложенное
                ws.cell(row=row_num, column=5, value=row[3])  # Цена в рублях
                ws.cell(row=row_num, column=6, value=f"=C{row_num}*E{row_num}")  # Сумма
                ws.cell(row=row_num, column=7, value=sheet_name)  # Поставщик
                ws.cell(row=row_num, column=8, value=row[4])  # Сроки
                ws.cell(row=row_num, column=9, value=payment_terms.get(sheet_name, ""))
                ws.cell(row=row_num, column=10, value=row[5])  # Комментарий
                
                row_num += 1
        
        # Форматирование ширины колонок
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 7  # Зафиксирована ширина 7
        ws.column_dimensions['C'].width = 7  # Зафиксирована ширина 7
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 40
        
        # Применяем валютный формат к колонкам E и F (цена и сумма в рублях)
        currency_format = '#,##0.00 ₽'
        for row in range(2, row_num):
            ws.cell(row=row, column=5).number_format = currency_format  # Колонка E
            ws.cell(row=row, column=6).number_format = currency_format  # Колонка F
        
        # Применяем тонкие границы ко всем заполненным ячейкам
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, row_num):
            for col in range(1, 11):  # Колонки A-J (1-10)
                ws.cell(row=row, column=col).border = thin_border
        
        return summary_wb
        
    except Exception as e:
        print(f"Ошибка при создании свода для одного товара: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


# Flask веб-приложение
app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# HTML шаблон для веб-интерфейса
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Сравниватель КП - Создание свода</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            max-width: 800px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f5f5f5;
        }
        .container {
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        h1 {
            color: #333;
            text-align: center;
            margin-bottom: 30px;
        }
        .upload-area {
            border: 2px dashed #ccc;
            border-radius: 10px;
            padding: 40px;
            text-align: center;
            margin: 20px 0;
            background-color: #fafafa;
        }
        .upload-area:hover {
            border-color: #007bff;
            background-color: #f0f8ff;
        }
        input[type="file"] {
            margin: 20px 0;
            padding: 10px;
            border: 1px solid #ddd;
            border-radius: 5px;
            width: 100%;
            max-width: 400px;
        }
        button {
            background-color: #007bff;
            color: white;
            padding: 12px 30px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-size: 16px;
            margin-top: 20px;
        }
        button:hover {
            background-color: #0056b3;
        }
        button:disabled {
            background-color: #ccc;
            cursor: not-allowed;
        }
        .alert {
            padding: 15px;
            margin: 20px 0;
            border-radius: 5px;
        }
        .alert-success {
            background-color: #d4edda;
            color: #155724;
            border: 1px solid #c3e6cb;
        }
        .alert-error {
            background-color: #f8d7da;
            color: #721c24;
            border: 1px solid #f5c6cb;
        }
        .instructions {
            background-color: #e9ecef;
            padding: 20px;
            border-radius: 5px;
            margin: 20px 0;
        }
        .instructions h3 {
            margin-top: 0;
            color: #495057;
        }
        .instructions ul {
            margin: 10px 0;
            padding-left: 20px;
        }
        .instructions li {
            margin: 5px 0;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>🔄 Сравниватель КП</h1>
        <h2 style="text-align: center; color: #666;">Создание сводной таблицы</h2>
        
        <div class="instructions">
            <h3>📋 Инструкция:</h3>
            <ul>
                <li>Загрузите Excel файл с коммерческими предложениями</li>
                <li>Файл должен содержать несколько листов (первый лист игнорируется)</li>
                <li>Каждый лист должен представлять предложение от одного поставщика</li>
                <li>Структура: Наименование | Количество | Количество предложенное | Цена | Сроки | Комментарий</li>
                <li>После обработки автоматически скачается сводная таблица</li>
            </ul>
        </div>

        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, message in messages %}
                    <div class="alert alert-{{ 'success' if category == 'success' else 'error' }}">
                        {{ message }}
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <form method="post" enctype="multipart/form-data">
            <div class="upload-area">
                <h3>📁 Выберите Excel файл</h3>
                <input type="file" name="file" accept=".xlsx,.xls" required>
                <br>
                <button type="submit">🚀 Создать сводную таблицу</button>
            </div>
        </form>
    </div>

    <script>
        // Добавляем интерактивность
        const fileInput = document.querySelector('input[type="file"]');
        const button = document.querySelector('button');
        
        fileInput.addEventListener('change', function() {
            if (this.files.length > 0) {
                button.textContent = '🚀 Обработать файл: ' + this.files[0].name;
            } else {
                button.textContent = '🚀 Создать сводную таблицу';
            }
        });
        
        // Показываем прогресс при отправке
        document.querySelector('form').addEventListener('submit', function() {
            button.disabled = true;
            button.textContent = '⏳ Обработка файла...';
        });
    </script>
</body>
</html>
'''

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # Проверяем, был ли загружен файл
        if 'file' not in request.files:
            flash('Файл не выбран', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        
        # Проверяем, что файл выбран
        if file.filename == '':
            flash('Файл не выбран', 'error')
            return redirect(request.url)
        
        # Проверяем расширение файла
        if not allowed_file(file.filename):
            flash('Неподдерживаемый формат файла. Используйте .xlsx или .xls', 'error')
            return redirect(request.url)
        
        temp_input_path = None
        temp_output_path = None
        
        try:
            # Создаем временный файл для загруженного файла
            temp_input = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_input_path = temp_input.name
            temp_input.close()
            file.save(temp_input_path)
            
            # Обрабатываем файл с помощью существующей функции
            summary_wb = build_summary_table(temp_input_path)
            if not summary_wb:
                flash('Ошибка при обработке файла', 'error')
                return redirect(request.url)
            
            # Создаем временный файл для результата
            temp_output = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_output_path = temp_output.name
            temp_output.close()
            
            # Сохраняем результат
            summary_wb.save(temp_output_path)
            
            # Генерируем имя для скачиваемого файла
            original_name = secure_filename(file.filename)
            base_name = os.path.splitext(original_name)[0]
            output_filename = f"{base_name}_свод.xlsx"
            
            # Отправляем файл пользователю
            response = send_file(
                temp_output_path,
                as_attachment=True,
                download_name=output_filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Удаляем временные файлы после отправки
            def cleanup():
                if temp_input_path and os.path.exists(temp_input_path):
                    os.unlink(temp_input_path)
                if temp_output_path and os.path.exists(temp_output_path):
                    os.unlink(temp_output_path)
            
            response.call_on_close(cleanup)
            return response
            
        except Exception as e:
            # Очищаем временные файлы в случае ошибки
            if temp_input_path and os.path.exists(temp_input_path):
                os.unlink(temp_input_path)
            if temp_output_path and os.path.exists(temp_output_path):
                os.unlink(temp_output_path)
                
            flash(f'Ошибка при обработке файла: {str(e)}', 'error')
            return redirect(request.url)
    
    return render_template_string(HTML_TEMPLATE)

if __name__ == "__main__":
    print("Запуск веб-приложения Сравниватель КП...")
    print("Откройте в браузере: http://localhost:5000")
    print("Для остановки нажмите Ctrl+C")
    app.run(debug=True, host='0.0.0.0', port=5000)
